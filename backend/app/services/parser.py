"""
File Parser Service module for SecureData Web.
Parses CSV and Excel (XLSX) files transiently to retrieve column headers
and sample rows for previews without persisting files to disk.
"""

import io
import pandas as pd
from openpyxl import load_workbook

def parse_csv_preview(file_buffer: io.BytesIO) -> tuple[list[str], list[list]]:
    """
    Parses the first 3 rows of a CSV file buffer to construct preview structures.
    Uses UTF-8 encoding by default, falling back to Latin-1 if decode fails.
    Args:
        file_buffer (io.BytesIO): In-memory buffer of the uploaded CSV.
    Returns:
        tuple[list[str], list[list]]: A tuple containing headers list and row values list.
    """
    file_buffer.seek(0)
    try:
        df = pd.read_csv(file_buffer, nrows=3, dtype=str)
    except UnicodeDecodeError:
        file_buffer.seek(0)
        df = pd.read_csv(file_buffer, nrows=3, dtype=str, encoding='latin1')
    
    df = df.fillna("")
    headers = df.columns.tolist()
    rows = df.values.tolist()
    return headers, rows

def parse_xlsx_preview(file_buffer: io.BytesIO) -> tuple[list[str], list[list]]:
    """
    Parses the active sheet of an Excel (.xlsx) file buffer using openpyxl.
    Retrieves the first row as headers and up to the next 3 rows as preview data.
    Args:
        file_buffer (io.BytesIO): In-memory buffer of the uploaded Excel file.
    Returns:
        tuple[list[str], list[list]]: A tuple containing headers list and row values list.
    """
    file_buffer.seek(0)
    wb = load_workbook(file_buffer, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = []
        rows = []
        
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                headers = [str(cell) if cell is not None else "" for cell in row]
            else:
                row_data = [str(cell) if cell is not None else "" for cell in row]
                rows.append(row_data)
            
            if i == 4:
                break
        return headers, rows
    finally:
        wb.close()
